import numpy as np 
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import psycopg2
from sqlalchemy import create_engine, text
from jinja2 import Template
import pdfkit
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import base64
import os
import re
from datetime import datetime
from collections import defaultdict

import geopandas as gpd
import plotly.express as px
import json
import shapely.wkt
import shapely.geometry
import io
from io import BytesIO
import base64
import markdown
from dotenv.main import load_dotenv
from openai import OpenAI

import plotly.graph_objects as go
import logging
import traceback
from fastapi import APIRouter, HTTPException
from fastapi.responses import HTMLResponse

from sqlalchemy import text
from fastapi import Depends
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from controller.db import SessionLocal, get_db

from routes.crop_annualburnprobability import create_combined_map
import concurrent.futures


# Configure basic logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Utility"], prefix='/Utility')

load_dotenv('fastApi/.env')

# client = OpenAI(api_key=os.environ['openai_api_key'])
client = None

# County abbreviation to full name mapping
COUNTY_MAPPING = {
    'ALA': 'Alameda', 'ALP': 'Alpine', 'AMA': 'Amador', 'BUT': 'Butte',
    'CAL': 'Calaveras', 'CC': 'Contra Costa', 'COL': 'Colusa', 'DN': 'Del Norte',
    'ED': 'El Dorado', 'FRE': 'Fresno', 'GLE': 'Glenn', 'HUM': 'Humboldt',
    'IMP': 'Imperial', 'INY': 'Inyo', 'KER': 'Kern', 'KIN': 'Kings',
    'LA': 'Los Angeles', 'LAK': 'Lake', 'LAS': 'Lassen', 'MAD': 'Madera',
    'MEN': 'Mendocino', 'MER': 'Merced', 'MNO': 'Mono', 'MOD': 'Modoc',
    'MON': 'Monterey', 'MPA': 'Mariposa', 'MRN': 'Marin', 'NAP': 'Napa',
    'NEV': 'Nevada', 'ORA': 'Orange', 'PLA': 'Placer', 'PLU': 'Plumas',
    'RIV': 'Riverside', 'SAC': 'Sacramento', 'SB': 'Santa Barbara', 'SBD': 'San Bernardino',
    'SBT': 'San Benito', 'SCL': 'Santa Clara', 'SCR': 'Santa Cruz', 'SD': 'San Diego',
    'SF': 'San Francisco', 'SHA': 'Shasta', 'SIE': 'Sierra', 'SIS': 'Siskiyou',
    'SJ': 'San Joaquin', 'SLO': 'San Luis Obispo', 'SM': 'San Mateo', 'SOL': 'Solano',
    'SON': 'Sonoma', 'STA': 'Stanislaus', 'SUT': 'Sutter', 'TEH': 'Tehama',
    'TRI': 'Trinity', 'TUL': 'Tulare', 'TUO': 'Tuolumne', 'VEN': 'Ventura',
    'YOL': 'Yolo', 'YUB': 'Yuba'
}

# Agency abbreviation to full name mapping
AGENCY_MAPPING = {
    'CALSTA': 'California State Transportation Agency',
    'CA State Parks': 'California State Parks',
    'CNRA': 'California Natural Resources Agency',
    'DOD': 'Department of Defense',
    'DOI': 'Department of the Interior',
    'NPS': 'National Park Service',
    'OTHER': 'Other Agencies',
    'TIMBER': 'Timber Operations',
    'Timber Companies': 'Timber Companies',
    'USDA': 'United States Department of Agriculture',
    'US Department of Energy': 'United States Department of Energy'
}

# Activity category to full name mapping
ACTIVITY_CAT_MAPPING = {
    'GRAZING': 'Grazing',
    'LAND_PROTEC': 'Land Protection',
    'MECH_HFR': 'Mechanical & Hand Fuels Reduction',
    'NOT_DEFINED': 'Not Defined',
    'PRESCRIBED_FIRE': 'Prescribed Fire',
    'TIMB_HARV': 'Timber Harvest',
    'TREE_PLNTING': 'Tree Planting'
}

# Activity description to full name mapping
ACTIVITY_DESC_MAPPING = {
    "AMW_AREA_RESTOR": "Area-Wide Restoration",
    "BIOMASS_REMOVAL": "Biomass Removal",
    "BROADCAST_BURN": "Broadcast Burn",
    "CHAIN_CRUSH": "Chain Crushing",
    "CHIPPING": "Chipping",
    "CLEARCUT": "Clearcutting",
    "COMM_THIN": "Commercial Thinning",
    "CONVERSION": "Land Conversion",
    "DISCING": "Discing",
    "DOZER_LINE": "Dozer Line Construction",
    "EASEMENT": "Easement",
    "ECO_HAB_RESTORATION": "Ecological Habitat Restoration",
    "EROSION_CONTROL": "Erosion Control",
    "FEE_TITLE": "Fee Title Acquisition",
    "GRP_SELECTION_HARVEST": "Group Selection Harvest",
    "HABITAT_REVEG": "Habitat Revegetation",
    "HANDLINE": "Handline Construction",
    "HERBICIDE_APP": "Herbicide Application",
    "INV_PLANT_REMOVAL": "Invasive Plant Removal",
    "LAND_ACQ": "Land Acquisition",
    "LANDING_TRT": "Landing Treatment",
    "LOP_AND_SCAT": "Lop and Scatter",
    'MASTICATION': 'Mastication',
    'MOWING': 'Mowing',
    'NOT_DEFINED': 'Not Defined',
    'OAK_WDLND_MGMT': 'Oak Woodland Management',
    'PEST_CNTRL': 'Pest Control',
    'PILE_BURN': 'Pile Burning',
    'PILING': 'Piling',
    'PL_TREAT_BURNED': 'Planned Treatment (Burned)',
    'PRESCRB_HERBIVORY': 'Prescribed Herbivory',
    'PRUNING': 'Pruning',
    'REHAB_UNDRSTK_AREA': 'Rehabilitation Understocked Area',
    'ROAD_CLEAR': 'Road Clearing',
    'ROAD_OBLITERATION': 'Road Obliteration',
    'SALVG_HARVEST': 'Salvage Harvest',
    'SANI_HARVEST': 'Sanitation Harvest',
    'SEEDBED_PREP': 'Seedbed Preparation',
    'SHELTERWD_REM_STEP': 'Shelterwood Removal Step',
    'SINGLE_TREE_SELECTION': 'Single Tree Selection',
    'SITE_PREP': 'Site Preparation',
    'SLASH_DISPOSAL': 'Slash Disposal',
    'SP_PRODUCTS': 'Special Products',
    'STREAM_CHNL_IMPRV': 'Stream Channel Improvement',
    'TBD': 'To Be Determined',
    'THIN_MAN': 'Thinning (Manual)',
    'THIN_MECH': 'Thinning (Mechanical)',
    'TRANSITION_HARVEST': 'Transition Harvest',
    'TREE_FELL': 'Tree Felling',
    'TREE_PLNTING': 'Tree Planting',
    'TREE_RELEASE_WEED': 'Tree Release/Weeding',
    'TREE_SEEDING': 'Tree Seeding',
    'UTIL_RIGHTOFWAY_CLR': 'Utility Right-of-Way Clearing',
    'VARIABLE_RETEN_HARVEST': 'Variable Retention Harvest',
    'WETLAND_RESTOR': 'Wetland Restoration',
    'WM_RESRC_BENEFIT': 'Watershed Management Resource Benefit',
    'YARDING': 'Yarding'
}

# Vegetation type to full name mapping
VEGETATION_MAPPING = {
    'AGRICULTURE': 'Agricultural Land',
    'FOREST': 'Forest Land',
    'GRASS_HERB': 'Grassland/Herbaceous',
    'SHRB_CHAP': 'Shrub/Chaparral',
    'SPARSE': 'Sparse Vegetation',
    'Trees Removed': 'Trees Removed',
    'URBAN': 'Urban/Developed',
    'WATER': 'Water',
    'WETLAND': 'Wetland'
}

# Ownership group to full name mapping
OWNERSHIP_MAPPING = {
    'FEDERAL': 'Federal Government',
    'LOCAL': 'Local Government',
    'NGO': 'Non-Governmental Organization',
    'PRIVATE_INDUSTRY': 'Private Industry',
    'PRIVATE_NON-INDUSTRY': 'Private Non-Industry',
    'STATE': 'State Government',
    'TRIBAL': 'Tribal Land'
}

# Region to full name mapping
REGION_MAPPING = {
    'CENTRAL_COAST': 'Central Coast',
    'NORTH_COAST': 'North Coast',
    'SIERRA_NEVADA': 'Sierra Nevada',
    'SOUTHERN_CA': 'Southern California',
    'Non-Spatial Data': 'Non-Spatial Data'
}

# Activity status to full name mapping
STATUS_MAPPING = {
    'ACTIVE': 'Active',
    'CANCELLED': 'Cancelled',
    'COMPLETE': 'Complete',
    'PLANNED': 'Planned'
}

# Administering organization to full name mapping
ADMIN_ORG_MAPPING = {
    'BIA': 'Bureau of Indian Affairs',
    'BLM': 'Bureau of Land Management',
    'BOF': 'Board of Forestry',
    'CALFIRE': 'California Department of Forestry and Fire Protection',
    'CALTRANS': 'California Department of Transportation',
    'CCC': 'California Conservation Corps',
    'CDFW': 'California Department of Fish and Wildlife',
    'DOC': 'Department of Conservation',
    'DOD': 'Department of Defense',
    'FWS': 'Fish and Wildlife Service',
    'MRCA': 'Mountains Recreation and Conservation Authority',
    'NPS': 'National Park Service',
    'OTHER': 'Other Organizations',
    'PARKS': 'California State Parks',
    'RMC': 'Rivers and Mountains Conservancy',
    'SCC': 'State Coastal Conservancy',
    'SDRC': 'San Diego River Conservancy',
    'SMMC': 'Santa Monica Mountains Conservancy',
    'SNC': 'Sierra Nevada Conservancy',
    'TAHOE': 'Tahoe Conservancy',
    'TIMBER': 'Timber Operations',
    'US Department of Energy': 'United States Department of Energy',
    'USFS': 'United States Forest Service',
    'WCB': 'Wildlife Conservation Board'
}


def create_db_connection():
    try:
        connection_string = os.environ['database_connection']
        engine = create_engine(connection_string)
        return engine
    except Exception as e:
        print(f"Error connecting to the database: {e}")
        return None

def get_county_data(db, county_code):
    query = f"""
    SELECT * FROM its.activities_report_20241209
    WHERE county = '{county_code}'
    AND year_txt ~ '^[0-9]+$' 
    AND CAST(year_txt AS INTEGER) BETWEEN 2021 AND 2023
    """
    with db.bind.connect() as conn:
        return pd.read_sql_query(text(query), conn)

def plot_to_base64(plt_figure):
    buf = io.BytesIO()
    plt_figure.savefig(buf, format='png', dpi=300, bbox_inches='tight')
    buf.seek(0)
    img_str = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close(plt_figure)
    return img_str

def adjust_chart_scaling(ax, data):
    max_val = data.max().max()
    min_val = data.min().min()
    if min_val == 0:
        min_val = 1e-6
    if False and max_val / min_val > 100:
        ax.set_yscale("log")
        ax.set_ylabel("Treatment Acres (Log Scale)")
    else:
        ax.set_ylabel("Treatment Acres")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

def create_agency_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["agency_name"] = df["agency"].map(AGENCY_MAPPING).fillna(df["agency"])
    agency_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "agency_name"])['activity_quantity'].sum().unstack(fill_value=0)
    agency_year_acres = agency_year_acres.clip(upper=np.percentile(agency_year_acres, 95))
    fig, ax = plt.subplots(figsize=(12, 6))
    agency_year_acres.plot(kind="bar", ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Agency (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, agency_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Agency", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_activity_cat_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["activity_cat_name"] = df["activity_cat"].map(ACTIVITY_CAT_MAPPING).fillna(df["activity_cat"])
    activity_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "activity_cat_name"])['activity_quantity'].sum().unstack(fill_value=0)
    activity_year_acres = activity_year_acres.clip(upper=np.percentile(activity_year_acres, 95))
    fig, ax = plt.subplots(figsize=(10, 6))
    activity_year_acres.plot(kind="bar", ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Category (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, activity_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Category", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_vegetation_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["vegetation_name"] = df["broad_vegetation_type"].map(VEGETATION_MAPPING).fillna(df["broad_vegetation_type"])
    veg_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "vegetation_name"])['activity_quantity'].sum().unstack(fill_value=0)
    veg_year_acres = veg_year_acres.clip(upper=np.percentile(veg_year_acres, 95))
    fig, ax = plt.subplots(figsize=(10, 6))
    veg_year_acres.plot(kind="bar", ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Vegetation Type (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, veg_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Vegetation Type", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_ownership_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["ownership_name"] = df["primary_ownership_group"].map(OWNERSHIP_MAPPING).fillna(df["primary_ownership_group"])
    ownership_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "ownership_name"])['activity_quantity'].sum().unstack(fill_value=0)
    ownership_year_acres = ownership_year_acres.clip(upper=np.percentile(ownership_year_acres, 95))
    fig, ax = plt.subplots(figsize=(10, 6))
    ownership_year_acres.plot(kind="bar", ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Land Ownership (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, ownership_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Ownership", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_status_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["status_name"] = df["activity_status"].map(STATUS_MAPPING).fillna(df["activity_status"])
    status_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "status_name"])['activity_quantity'].sum().unstack(fill_value=0)
    status_year_acres = status_year_acres.clip(upper=np.percentile(status_year_acres, 95))
    fig, ax = plt.subplots(figsize=(10, 6))
    status_year_acres.plot(kind="bar", stacked=True, ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Status (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, status_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Status", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_admin_org_chart(df):
    plt.style.use('seaborn-v0_8-whitegrid')
    df['admin_org_name'] = df['administering_org'].map(ADMIN_ORG_MAPPING).fillna(df['administering_org'])
    top_orgs = df[df['activity_uom'] == 'AC'].groupby('admin_org_name')['activity_quantity'].sum().nlargest(10)
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = plt.cm.viridis(np.linspace(0.1, 0.9, len(top_orgs)))
    bars = ax.barh(top_orgs.index[::-1], top_orgs.values[::-1], color=colors[::-1], height=0.7, edgecolor='none')
    for i, (value, bar) in enumerate(zip(top_orgs.values[::-1], bars)):
        ax.text(value + (value * 0.02), i, f'{int(value):,}', va='center', ha='left', fontsize=10, fontweight='bold')
    ax.set_title('Top 10 Administering Organizations by Acres', fontsize=14, fontweight='bold', pad=20)
    ax.set_xlabel('Acres Treated', fontsize=12)
    ax.set_ylabel('')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.3)
    plt.tight_layout()
    return fig

def generate_summary_statistics(df, county_name):
    df_filtered = df[df['activity_uom'] == 'AC']
    total_activities = len(df_filtered)
    total_acres = df_filtered['activity_quantity'].sum()
    years_active = sorted(df_filtered[df_filtered['year_txt'].str.isnumeric()]['year_txt'].astype(int).unique())
    top_activity = df_filtered.groupby('activity_description')['activity_quantity'].sum().idxmax()
    top_activity = ACTIVITY_DESC_MAPPING.get(top_activity, top_activity)
    main_agency = df_filtered.groupby('agency')['activity_quantity'].sum().idxmax()
    main_agency = AGENCY_MAPPING.get(main_agency, main_agency)
    return {
        'county_name': county_name,
        'total_activities': total_activities,
        'total_acres': f"{total_acres:,.0f}",
        'years_active': f"{min(years_active)} to {max(years_active)}",
        'top_activity': top_activity,
        'main_agency': main_agency
    }

def create_activity_description_table(df):
    df_filtered = df[df['activity_uom'] == 'AC'].copy()
    df_filtered['activity_desc_name'] = df_filtered['activity_description'].map(ACTIVITY_DESC_MAPPING).fillna(df_filtered['activity_description'])
    activity_acres = df_filtered.groupby('activity_desc_name')['activity_quantity'].sum().reset_index()
    activity_acres.columns = ['Treatment Activity', 'Total Acres Treated']
    total_acres = activity_acres['Total Acres Treated'].sum()
    activity_acres['Percentage'] = round(activity_acres['Total Acres Treated'] / total_acres * 100, 1)
    activity_acres['Percentage'] = activity_acres['Percentage'].astype(str) + '%'
    activity_acres['Total Acres Treated'] = activity_acres['Total Acres Treated'].map(lambda x: f"{x:,.2f}")
    html_table = activity_acres.to_html(index=False, classes='table table-striped table-hover', border=0)
    return html_table


def create_map_plot(db: Session, county_code):
    query = f"""
        SELECT 
            county, 
            ST_X(ST_Transform(ST_Centroid(geom), 4326)) as lon,
            ST_Y(ST_Transform(ST_Centroid(geom), 4326)) as lat
        FROM its.activities_report_20241209
        WHERE county = '{county_code}'
    """
    with db.bind.connect() as conn:
        df = pd.read_sql_query(text(query), conn)
    if df.empty:
        return None
    fig = go.Figure(go.Scattermapbox(
        mode="markers",
        lon=df['lon'],
        lat=df['lat'],
        marker=dict(size=5, color='blue'),
        hoverinfo='none'
    ))
    fig.update_layout(
        mapbox_style="open-street-map",
        mapbox=dict(center=dict(lat=df['lat'].mean(), lon=df['lon'].mean()), zoom=7),
        margin=dict(l=5, r=5, t=5, b=5),
        height=400,
        width=600
    )
    img_bytes = io.BytesIO()
    fig.write_image(img_bytes, format='png', engine='kaleido')
    img_bytes.seek(0)
    return base64.b64encode(img_bytes.read()).decode("utf-8")


def extract_data_insights(df, county_name):
    df_filtered = df[df['activity_uom'] == 'AC'].copy()
    yearly_totals = df_filtered.groupby('year_txt')['activity_quantity'].sum()
    try:
        peak_year = yearly_totals.idxmax()
        peak_amount = yearly_totals.max()
    except:
        peak_year = "N/A"
        peak_amount = 0
    if len(yearly_totals) > 1:
        years = sorted(yearly_totals.index)
        first_year = years[0]
        last_year = years[-1]
        first_amount = yearly_totals[first_year]
        last_amount = yearly_totals[last_year]
        percent_change = ((last_amount - first_amount) / first_amount * 100) if first_amount > 0 else 0
        trend_direction = "increased" if percent_change > 0 else "decreased"
    else:
        percent_change = 0
        trend_direction = "remained stable"
    top_activities = df_filtered.groupby('activity_description')['activity_quantity'].sum().nlargest(3)
    top_activities_names = [ACTIVITY_DESC_MAPPING.get(act, act) for act in top_activities.index]
    top_veg_types = df_filtered.groupby('broad_vegetation_type')['activity_quantity'].sum().nlargest(3)
    top_veg_names = [VEGETATION_MAPPING.get(veg, veg) for veg in top_veg_types.index]
    top_ownership = df_filtered.groupby('primary_ownership_group')['activity_quantity'].sum().nlargest(3)
    top_ownership_names = [OWNERSHIP_MAPPING.get(own, own) for own in top_ownership.index]
    status_dist = df_filtered.groupby('activity_status')['activity_quantity'].sum()
    total_acres = status_dist.sum()
    complete_pct = (status_dist.get('COMPLETE', 0) / total_acres * 100) if total_acres > 0 else 0
    planned_pct = (status_dist.get('PLANNED', 0) / total_acres * 100) if total_acres > 0 else 0
    agency_contrib = df_filtered.groupby('agency')['activity_quantity'].sum().nlargest(2)
    top_agencies = [AGENCY_MAPPING.get(ag, ag) for ag in agency_contrib.index]
    insights = {
        "county_name": county_name,
        "total_acres_treated": df_filtered['activity_quantity'].sum(),
        "peak_year": peak_year,
        "peak_amount": peak_amount,
        "trend_direction": trend_direction,
        "percent_change": round(abs(percent_change), 1),
        "top_activities": top_activities_names,
        "top_vegetation_types": top_veg_names,
        "top_ownership_types": top_ownership_names,
        "completed_treatments_percent": round(complete_pct, 1),
        "planned_treatments_percent": round(planned_pct, 1),
        "lead_agencies": top_agencies
    }
    return insights

# Function to generate a prompt for the LLM
def create_llm_prompt(insights):
    """Create a detailed prompt for the LLM based on data insights."""
    prompt = f"""
You are an expert of Wildfire & Landscape Resilience Interagency Treatment Dashboard. The following is the description of the Dashboard:

As California works toward ambitious wildfire and landscape resilience goals, transparency and effective planning tools are critical to success. The California Wildfire and Landscape Interagency Treatment Dashboard, for the first time ever in California, provides a single source for displaying recently completed forest and wildland projects from over a dozen different federal and state agencies.

The Dashboard is a first-of-its-kind platform that displays the location and size of federal and state wildfire and landscape resilience treatments throughout the state. The Dashboard is a highly interactive online tool by which users can sort treatments by region, county, land ownership, and more. By charting the work of what has been accomplished to date, the Dashboard can be used to guide practitioners on where to plan new projects.
 
You're generating content for an official report about treatment activities in {insights['county_name']} County, California.

Based on the following data points:
- Total acres treated from 2021-2023: {insights['total_acres_treated']:,.2f} acres
- The peak treatment year was {insights['peak_year']} with {insights['peak_amount']:,.2f} acres
- Treatment acres have {insights['trend_direction']} by {insights['percent_change']}% from first to last year
- Top treatment activities: {', '.join(insights['top_activities'])}
- Main vegetation types treated: {', '.join(insights['top_vegetation_types'])}
- Primary land ownership where treatments occurred: {', '.join(insights['top_ownership_types'])}
- Treatments that are complete: {insights['completed_treatments_percent']}%
- Treatments that are still planned: {insights['planned_treatments_percent']}%
- Lead agencies: {', '.join(insights['lead_agencies'])}

Please create three sections for our report in the following format:
1. EXECUTIVE SUMMARY: A detailed 2-paragraph executive summary that highlights the most significant findings from this data.

2. KEY FINDINGS: 4-5 short paragraphs analyzing the data, connecting treatment patterns to potential wildfire risk reduction and landscape resilience, noting trends, and highlighting important relationships between agencies, land types, and treatment methods.

3. RECOMMENDATIONS: 4-5 specific, actionable recommendations based on the data, including suggestions for future treatment priorities, interagency cooperation opportunities, and potential gaps that need addressing.

Use an authoritative, professional tone. Be factual and specific, avoiding vague generalizations. Mention specific agencies, vegetation types, and treatment methods where relevant. Make your recommendations specific and actionable. 

You can use markdown in these sections.
"""
    return prompt


def generate_llm_content(prompt):
    try:
       response = client.chat.completions.create(
            model="gpt-4-turbo",
            # model="gpt-4o-mini", 
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7)
       return response.choices[0].message.content
    except Exception as e:
       return f"AI service error: {str(e)}"

def generate_report_content(insights):
    prompt = create_llm_prompt(insights)
    llm_content = generate_llm_content(prompt)
    
    # More flexible regex patterns that handle variations in formatting
    executive_summary = re.search(r"1\.\s*EXECUTIVE\s*SUMMARY\s*\n(.*?)(?=\s*2\.\s*KEY\s*FINDINGS)", llm_content, re.DOTALL)
    key_findings = re.search(r"2\.\s*KEY\s*FINDINGS\s*\n(.*?)(?=\s*3\.\s*RECOMMENDATIONS)", llm_content, re.DOTALL)
    recommendations = re.search(r"3\.\s*RECOMMENDATIONS\s*\n(.*?)(?=$)", llm_content, re.DOTALL)
    
    # Fallback patterns if the above don't match
    if not executive_summary:
        executive_summary = re.search(r"EXECUTIVE\s*SUMMARY[:\s]*(.*?)(?=\s*KEY\s*FINDINGS|RECOMMENDATIONS|$)", llm_content, re.DOTALL)
    
    if not key_findings:
        key_findings = re.search(r"KEY\s*FINDINGS[:\s]*(.*?)(?=\s*RECOMMENDATIONS|EXECUTIVE\s*SUMMARY|$)", llm_content, re.DOTALL)
    
    if not recommendations:
        recommendations = re.search(r"RECOMMENDATIONS[:\s]*(.*?)(?=$)", llm_content, re.DOTALL)
    
    # Extract and clean the sections
    executive_summary_text = executive_summary.group(1).strip() if executive_summary else "Not found"
    key_findings_text = key_findings.group(1).strip() if key_findings else "Not found"
    recommendations_text = recommendations.group(1).strip() if recommendations else "Not found"
    
    # Clean up trailing section numbers
    executive_summary_text = re.sub(r'\s+\d+\.\s*$', '', executive_summary_text)
    key_findings_text = re.sub(r'\s+\d+\.\s*$', '', key_findings_text)
    
    return {
        "executive_summary": markdown.markdown(executive_summary_text),
        "key_findings": markdown.markdown(key_findings_text),
        "recommendations": markdown.markdown(recommendations_text)
    }


def generate_llm_content_concurrent(insights):
    return generate_report_content(insights)

def create_overlay_map_concurrent(db, county_code):
    return create_combined_map(
        db=db,
        county_name=COUNTY_MAPPING[county_code],
        geoserver_url="https://sparcal.sdsc.edu/geoserver",
        layer_name="rrk:annualburnprobability_202212_202406_t1_v5",
        output_png=None,
        county_code=county_code,
        title=f"",
        colormap='YlOrRd',
        point_color='blue',
        point_size=15,
        point_alpha=0.7
    )


def create_admin_org_table(df):
    """Create styled table for administering org data using activity categories"""
    admin_df = df[df['activity_uom'] == 'AC'].groupby(
        ['administering_org', 'activity_cat', 'year_txt']
    )['activity_quantity'].sum().reset_index()
    
    admin_df['Administering Organization'] = admin_df['administering_org'].map(ADMIN_ORG_MAPPING)
    admin_df['Activity Category'] = admin_df['activity_cat'].map(ACTIVITY_CAT_MAPPING)
    admin_df['Year'] = admin_df['year_txt']
    admin_df['Acres'] = admin_df['activity_quantity'].apply(lambda x: f"{x:,.2f}")
    
    return admin_df[['Administering Organization', 'Activity Category', 'Year', 'Acres']] \
        .sort_values(['Year', 'Administering Organization']) \
        .to_html(index=False, classes='table table-striped table-hover', border=0)

def create_ownership_table(df):
    """Create styled table for ownership data using activity categories"""
    ownership_df = df[df['activity_uom'] == 'AC'].groupby(
        ['primary_ownership_group', 'activity_cat', 'year_txt']
    )['activity_quantity'].sum().reset_index()
    
    ownership_df['Ownership'] = ownership_df['primary_ownership_group'].map(OWNERSHIP_MAPPING)
    ownership_df['Activity Category'] = ownership_df['activity_cat'].map(ACTIVITY_CAT_MAPPING)
    ownership_df['Year'] = ownership_df['year_txt']
    ownership_df['Acres'] = ownership_df['activity_quantity'].apply(lambda x: f"{x:,.2f}")
    
    return ownership_df[['Ownership', 'Activity Category', 'Year', 'Acres']] \
        .sort_values(['Year', 'Ownership']) \
        .to_html(index=False, classes='table table-striped table-hover', border=0)


# Add this helper function
def generate_excel_data(df, county_name):
    """Generate Excel data with proper multi-column sorting"""
    from openpyxl.utils import get_column_letter
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Helper function for column width adjustment
        def auto_adjust_columns(worksheet):
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # 1. Administering Organizations (sorted: Org → Category → Year)
        admin_df = df[df['activity_uom'] == 'AC'].groupby(
            ['administering_org', 'activity_cat', 'year_txt']
        )['activity_quantity'].sum().reset_index()
        
        admin_df['Administering Organization'] = admin_df['administering_org'].map(ADMIN_ORG_MAPPING)
        admin_df['Activity Category'] = admin_df['activity_cat'].map(ACTIVITY_CAT_MAPPING)
        admin_df['Year'] = admin_df['year_txt'].astype(int)
        admin_df['Acres'] = admin_df['activity_quantity']
        
        admin_df = admin_df[['Administering Organization', 'Activity Category', 'Year', 'Acres']] \
            .sort_values(['Administering Organization', 'Activity Category', 'Year'])  # Changed sort order
        
        admin_df.to_excel(writer, sheet_name='Administering Organizations', index=False)
        auto_adjust_columns(writer.sheets['Administering Organizations'])

        # 2. Land Ownership (sorted: Ownership → Category → Year)
        ownership_df = df[df['activity_uom'] == 'AC'].groupby(
            ['primary_ownership_group', 'activity_cat', 'year_txt']
        )['activity_quantity'].sum().reset_index()
        
        ownership_df['Ownership'] = ownership_df['primary_ownership_group'].map(OWNERSHIP_MAPPING)
        ownership_df['Activity Category'] = ownership_df['activity_cat'].map(ACTIVITY_CAT_MAPPING)
        ownership_df['Year'] = ownership_df['year_txt'].astype(int)
        ownership_df['Acres'] = ownership_df['activity_quantity']
        
        ownership_df = ownership_df[['Ownership', 'Activity Category', 'Year', 'Acres']] \
            .sort_values(['Ownership', 'Activity Category', 'Year'])  # Changed sort order
        
        ownership_df.to_excel(writer, sheet_name='Land Ownership', index=False)
        auto_adjust_columns(writer.sheets['Land Ownership'])

        # Format numeric columns
        for sheet in writer.sheets.values():
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 4:  # Acres column
                        cell.number_format = '#,##0.00'
                    elif cell.column == 3:  # Year column
                        cell.number_format = '0'

    output.seek(0)
    return output


@router.get("/its_report", include_in_schema=True)
def create_county_report(county_code, output_format='html', db: Session = Depends(get_db)):
    county_code = county_code.upper()
    if county_code not in COUNTY_MAPPING:
        return f"Error: '{county_code}' is not a valid California county code."
    
    county_name = COUNTY_MAPPING[county_code]
    engine = create_db_connection()
    if not engine:
        return "Error: Could not connect to the database."
    logger.info(f"Creating a report for {county_name}")
      
    try:
        df = get_county_data(db, county_code)
        if len(df) == 0:
            return f"No treatment data found for {county_name} County."
        logger.info(f"Treatment data for {county_name} County loaded: {df.shape[0]}")

        # Add Excel output handling
        if output_format.lower() == 'xlsx':
            excel_data = generate_excel_data(df, county_name)
            return StreamingResponse(
                excel_data,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={county_name}_treatment_report.xlsx"}
            )
        
        # Extract insights early for LLM content generation
        insights = extract_data_insights(df, county_name)

        # Generate charts while LLM and overlay map are being processed
        logger.info("Generating agency chart")
        agency_chart = plot_to_base64(create_agency_chart(df))
            
        logger.info("Generating activity category chart")
        activity_cat_chart = plot_to_base64(create_activity_cat_chart(df))

        logger.info("Generating vegetation type chart")
        vegetation_chart = plot_to_base64(create_vegetation_chart(df))

        logger.info("Generating ownership chart")
        ownership_chart = plot_to_base64(create_ownership_chart(df))

        logger.info("Generating status chart")
        status_chart = plot_to_base64(create_status_chart(df))

        logger.info("Generating admin chart")
        admin_org_chart = plot_to_base64(create_admin_org_chart(df))
            
        logger.info("Summary statistics and activity table")
        summary = generate_summary_statistics(df, county_name)
        activity_table = create_activity_description_table(df)

        logger.info("Generating overlay map")  
        # Generate overlay map
        try:
            overlay_map = create_combined_map(
                db=db,
                county_name=county_name,
                geoserver_url="https://sparcal.sdsc.edu/geoserver",
                layer_name="rrk:annualburnprobability_202212_202406_t1_v5",
                output_png=None,  # Not needed since we return base64
                county_code=county_code,
                title=f"",
                colormap='YlOrRd',
                point_color='blue',
                point_size=15,
                point_alpha=0.7
            )
            logger.info("Generated overlay map") 
        except Exception as e:
            traceback.print_exc()
            logger.error(f"Error generating overlay map: {str(e)}")
            overlay_map = None

        logger.info("LLM responded and overlay map generated")        

        logger.info("Generating spreadsheet tables")
        admin_table = create_admin_org_table(df)
        ownership_table = create_ownership_table(df)

        # Calculate statistics based on burn probability
        # high_risk_threshold = 0.05
        # high_risk_points = points_df[points_df['burn_probability'] > high_risk_threshold]
        # total_high_risk_acres = high_risk_points['activity_quantity'].sum()

        # Add high-risk acres to summary
        # summary['total_high_risk_acres'] = f"{total_high_risk_acres:,.2f}"
        # summary['high_risk_threshold'] = high_risk_threshold

        html_template = """
<!DOCTYPE html>
<html>
<head>
    <title>{{ summary.county_name }} County Wildfire & Landscape Resilience Report</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        :root {
            --primary-color: #1e5c97;
            --secondary-color: #2c8b57;
            --accent-color: #f39c12;
            --light-bg: #f8f9fa;
            --dark-bg: #2c3e50;
            --text-color: #333333;
            --light-text: #ffffff;
            --border-color: #e0e0e0;
            --box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }
        body {
            font-family: 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            line-height: 1.6;
            color: var(--text-color);
            max-width: 1200px;
            margin: 0 auto;
            padding: 0;
            background-color: #f5f7fa;
        }
        h1, h2, h3, h4 {
            font-weight: 600;
            color: var(--primary-color);
            margin-top: 1.5em;
            margin-bottom: 0.8em;
        }
        h1 {
            font-size: 2.5rem;
            margin-top: 0;
        }
        h2 {
            font-size: 1.8rem;
            border-bottom: 2px solid var(--border-color);
            padding-bottom: 0.3em;
        }
        h3 {
            font-size: 1.4rem;
            color: var(--secondary-color);
        }
        .container {
            background-color: white;
            box-shadow: var(--box-shadow);
            padding: 2rem;
            margin: 0 auto;
        }
        .header {
            background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
            color: var(--light-text);
            padding: 3rem;
            margin-bottom: 2rem;
            text-align: center;
            position: relative;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
        }
        .header h1, .header h2 {
            color: white;
            margin: 0.5rem 0;
            text-shadow: 1px 1px 3px rgba(0, 0, 0, 0.2);
        }
        .header p {
            margin-top: 1rem;
            font-size: 1.1rem;
            opacity: 0.9;
        }
        .section {
            margin-bottom: 2.5rem;
            padding: 0 1.5rem;
        }
        .introduction {
            background-color: var(--light-bg);
            border-left: 4px solid var(--primary-color);
            padding: 1.5rem;
            margin-bottom: 2rem;
            border-radius: 0 4px 4px 0;
        }
        .summary-box {
            background-color: white;
            border-radius: 8px;
            box-shadow: var(--box-shadow);
            padding: 1.5rem 2rem;
            margin-bottom: 2rem;
            border-top: 5px solid var(--accent-color);
        }
        .summary-box h2 {
            color: var(--accent-color);
            border-bottom: none;
            margin-top: 0;
        }
        .summary-stats {
            display: flex;
            flex-wrap: wrap;
            gap: 1.5rem;
            margin-top: 1.5rem;
        }
        .stat-item {
            flex: 1;
            min-width: 200px;
            background-color: var(--light-bg);
            padding: 1rem;
            border-radius: 6px;
            text-align: center;
        }
        .stat-value {
            font-size: 1.8rem;
            font-weight: 700;
            color: var(--primary-color);
            margin-bottom: 0.5rem;
        }
        .stat-label {
            font-size: 0.9rem;
            color: #666;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .chart-container {
            margin: 2rem 0;
        }
        .chart-row {
            display: flex;
            flex-wrap: wrap;
            gap: 2rem;
            margin-bottom: 2rem;
        }
        .chart {
            flex: 1;
            min-width: 300px;
            background-color: white;
            border-radius: 8px;
            box-shadow: var(--box-shadow);
            padding: 1.5rem;
            transition: transform 0.2s ease;
        }
        .chart:hover {
            transform: translateY(-5px);
        }
        .chart h3 {
            text-align: center;
            margin-top: 0;
            padding-bottom: 0.8rem;
            border-bottom: 1px solid var(--border-color);
        }
        .chart img {
            max-width: 100%;
            height: auto;
            display: block;
            margin: 1rem auto;
        }
        .chart-caption {
            font-size: 0.9rem;
            color: #666;
            text-align: center;
            margin-top: 1rem;
            font-style: italic;
        }
        .full-width {
            flex-basis: 100%;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin: 2rem 0;
            background-color: white;
            box-shadow: var(--box-shadow);
            border-radius: 8px;
            overflow: hidden;
        }
        th, td {
            padding: 12px 15px;
            text-align: left;
        }
        th {
            background-color: var(--primary-color);
            color: white;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.9rem;
            letter-spacing: 0.5px;
        }
        tr:nth-child(even) {
            background-color: #f2f7ff;
        }
        tr:hover {
            background-color: #e6f0ff;
        }
        .findings-box {
            background-color: var(--light-bg);
            padding: 1.5rem 2rem;
            border-radius: 8px;
            margin-bottom: 2rem;
            border-left: 4px solid var(--primary-color);
        }
        .recommendations {
            background-color: var(--light-bg);
            padding: 1.5rem 2rem;
            border-radius: 8px;
            border-left: 4px solid var(--secondary-color);
        }
        .recommendations ul {
            padding-left: 1.2rem;
        }
        .recommendations li {
            margin-bottom: 0.8rem;
        }
        .footer {
            margin-top: 3rem;
            padding: 2rem 0;
            text-align: center;
            font-size: 0.9rem;
            color: #777;
            border-top: 1px solid var(--border-color);
        }
        @media print {
            body {
                background-color: white;
            }
            .container {
                box-shadow: none;
                padding: 0;
            }
            .chart:hover {
                transform: none;
            }
            .chart-row {
                display: block;
            }
            .chart {
                width: 100%;
                margin-bottom: 2rem;
                box-shadow: none;
                page-break-inside: avoid;
            }
            .header {
                background: var(--primary-color) !important;
                -webkit-print-color-adjust: exact;
            }
            th {
                background-color: var(--primary-color) !important;
                color: white !important;
                -webkit-print-color-adjust: exact;
            }
            .summary-box {
                box-shadow: none;
                border: 1px solid var(--border-color);
            }
            table {
                box-shadow: none;
            }

    /* Chart sizing adjustments */
    .chart img {
        max-height: 320px !important;  /* Increased from 250px */
        width: auto !important;
        margin: 12px auto !important;
        page-break-inside: avoid;
    }

    /* Container adjustments */
    .chart {
        page-break-inside: avoid;
        margin: 8px 0 !important;
        padding: 4px !important;
    }

    /* Grid layout for charts */
    .chart-row {
        display: grid !important;
        grid-template-columns: 1fr 1fr;
        gap: 8px !important;
        page-break-inside: avoid;
    }

    /* Single chart full width */
    .full-width.chart img {
        max-height: 400px !important;
        width: 95% !important;
    }

        }


/* Chart image quality preservation */
.chart img {
    image-rendering: crisp-edges;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
}

/* PDF-specific image scaling */
@media print {
    canvas {
        max-width: 100% !important;
        height: auto !important;
    }
    
    figure {
        max-width: 90% !important;
        margin: 0 auto !important;
    }
}



    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <p>Wildfire & Landscape Resilience Task Force</p>
            <h2>Interagency Treatment Summary Report</h2>
            <p style="font-size:16pt; font-weight: bold;">{{ summary.county_name }} County </p>
        </div>
        
        <div class="section">
             <div class="summary-box">
                <h2>County Treatment Overview</h2>
                <div class="summary-stats">
                    <div class="stat-item">
                        <div class="stat-value">{{ summary.total_activities }}</div>
                        <div class="stat-label">Total Treatment Activities</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value">{{ summary.total_acres }}</div>
                        <div class="stat-label">Total Acres Treated</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value">{{ summary.years_active }}</div>
                        <div class="stat-label">Active Treatment Period</div>
                    </div>
                </div>
                <!--
                <div class="summary-stats" style="margin-top: 1rem;">
                    <div class="stat-item">
                        <div class="stat-value" style="font-size: 1.3rem;">{{ summary.top_activity }}</div>
                        <div class="stat-label">Primary Treatment Type</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value" style="font-size: 1.3rem;">{{ summary.main_agency }}</div>
                        <div class="stat-label">Lead Agency</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value">{{ summary.total_high_risk_acres }}</div>
                        <div class="stat-label">Acres Treated in High-Risk Areas (Burn Probability > {{ summary.high_risk_threshold }})</div>
                    </div>
                </div>
                -->
            </div>
        </div>
        
        <div class="section">
            <div class="chart-container">
               <div class="chart-row">
                    <div class="chart full-width">
                        <h3>Treatment Locations and Annual Burn Probability</h3>
                        {% if overlay_map %}
                            <img src="data:image/png;base64,{{ overlay_map }}" alt="Burn Probability Overlay">
                        {% else %}
                            <p>Overlay map could not be generated.</p>
                        {% endif %}
                        <!--
			<p class="chart-caption">Annual Burn Probability from the <a href="https://caregionalresourcekits.org/clm.html#fire_dyn" target="_blank">California Landscape Metrics</a> overlaid with treatment locations in {{ summary.county_name }} County</p>
                        -->
			<p class="table-description">
			    Annual Burn Probability - likelihood of a wildfire of any intensity occurring at a given location in a single fire season, <a href="https://caregionalresourcekits.org/clm.html#fire_dyn" target="_blank">California Landscape Metrics</a>, Pyrologix LLC, 2022
			</p>
                         
                   </div>
                </div>
             </div>
        </div>


        <div class="section">
            <h2>Treatment Activities Analysis</h2>
            <div class="chart-container">
                <div class="chart-row">
                    <div class="chart">
                        <h3>Treatment Categories Distribution</h3>
                        <img src="data:image/png;base64,{{ activity_cat_chart }}" alt="Treatment Categories">
                        <p class="chart-caption">Breakdown of treatment activities by category across 2021-2023.</p>
                    </div>
                    <div class="chart">
                        <h3>Vegetation Types Treated</h3>
                        <img src="data:image/png;base64,{{ vegetation_chart }}" alt="Vegetation Types Treated">
                        <p class="chart-caption">Distribution of treatments across vegetation types, 2021-2023.</p>
                    </div>
                </div>
                <div class="chart-row">
                    <div class="chart">
                        <h3>Land Ownership Distribution</h3>
                        <img src="data:image/png;base64,{{ ownership_chart }}" alt="Land Ownership">
                        <p class="chart-caption">Treatment distribution by land ownership, 2021-2023.</p>
                    </div>
                    <div class="chart">
                        <h3>Top Administering Organizations</h3>
                        <img src="data:image/png;base64,{{ admin_org_chart }}" alt="Top Administering Organizations">
                        <p class="chart-caption">Top 10 organizations administering treatments, 2021-2023.</p>
                    </div>
                </div>
                <!--
                <div class="chart-row">
                    <div class="chart">
                        <h3>Implementing Agencies</h3>
                        <img src="data:image/png;base64,{{ agency_chart }}" alt="Implementing Agencies">
                        <p class="chart-caption">Distribution of treatment activities by agency, 2021-2023.</p>
                    </div>
                    <div class="chart">
                        <h3>Current Treatment Status</h3>
                        <img src="data:image/png;base64,{{ status_chart }}" alt="Treatment Status">
                        <p class="chart-caption">Distribution of treatment statuses across 2021-2023.</p>
                    </div>
                </div>
                <div class="chart-row">
                    <div class="chart">
                        <h3>Geospatial Distribution of Treatments</h3>
                        <img src="data:image/png;base64,{{ map_chart }}" alt="Geospatial Map">
                        <p class="chart-caption">Spatial visualization of treatments in {{ summary.county_name }} County.</p>
                    </div>
                </div>
                <div class="chart-row">
                    <div class="chart full-width">
                        <h3>Treatment Locations and Annual Burn Probability</h3>
                        {% if overlay_map %}
                            <img src="data:image/png;base64,{{ overlay_map }}" alt="Burn Probability Overlay">
                        {% else %}
                            <p>Overlay map could not be generated.</p>
                        {% endif %}
			<p class="chart-caption">Annual Burn Probability from the <a href="https://caregionalresourcekits.org/clm.html#fire_dyn" target="_blank">California Landscape Metrics</a> overlaid with treatment locations in {{ summary.county_name }} County</p>
                    </div>
                 </div>
                 -->
            </div>
        </div>

        <div class="section">
            <h2>Detailed Treatment Data</h2>

             <h3>Administering Organizations</h3>
             <p class="table-description">
                   Breakdown of treatment acres by administering organization and activity type.
                   Data reflects actual completed treatments from 2021-2023.
             </p>
             {{ admin_table|safe }}

             <h3>Land Ownership Details</h3>
             <p class="table-description">
                   Distribution of treatments across land ownership types, showing annual 
                   implementation rates by activity category.
             </p>
             {{ ownership_table|safe }}

    </div>

        </div>

        
        <div class="footer">
            <p><strong>California Wildfire and Landscape Interagency Treatment Dashboard</strong></p>
            <p>Report Generated: {{ current_date }}</p>
        </div>
    </div>
</body>
</html>
        """
        logger.info("Creating HTML content")
        template = Template(html_template)
        html_content = template.render(
            summary=summary,
            activity_cat_chart=activity_cat_chart,
            vegetation_chart=vegetation_chart,
            status_chart=status_chart,
            ownership_chart=ownership_chart,
            agency_chart=agency_chart,
            admin_org_chart=admin_org_chart,
            activity_table=activity_table,
            # map_chart=map_chart,
            overlay_map=overlay_map,
            # report_content=report_content,
            admin_table=admin_table,
            ownership_table=ownership_table,
            current_date=datetime.now().strftime('%B %d, %Y')
        )
        logger.info("Created HTML content")
        if output_format.lower() == 'html':
            return HTMLResponse(content=html_content, status_code=200)
        elif output_format.lower() == 'pdf':
            options = {
                'page-size': 'Letter',
                'margin-top': '0.50in',
                'margin-right': '0.50in',
                'margin-bottom': '0.50in',
                'margin-left': '0.50in',
                'encoding': "UTF-8",
                'no-outline': None,
                'enable-local-file-access': None
            }

            options = {
                'page-size': 'Letter',
                'margin-top': '0.5in',  # Reduced from 0.75in
                'margin-right': '0.5in',
                'margin-bottom': '0.5in',
                'margin-left': '0.5in',
                'encoding': "UTF-8",
                'no-outline': None,
                'viewport-size': '1280x1024',
                'enable-local-file-access': None,
                'disable-smart-shrinking': None,  # Prevent automatic scaling
                'header-spacing': '0',  # Remove header spacing
                'footer-spacing': '0',  # Remove footer spacing
                'dpi': 400,  # Higher resolution for better element fitting
                'zoom': '1.0',  # Slightly shrink content to fit better
                'print-media-type': '',  # Use print CSS styles
                'user-style-sheet': 'pdf_styles.css'  # Optional CSS file for print optimization
            }


            pdf = pdfkit.from_string(html_content, False, options=options)
            return Response(content=pdf, media_type="application/pdf")
        else:
            return "Error: Output format must be either 'html' or 'pdf'."
    except Exception as e:
        return f"Error generating report: {str(e)}"

if __name__ == "__main__":
    # Example usage:
    # html_report = create_county_report('LA', 'html')
    # with open('los_angeles_report.html', 'w', encoding='utf-8') as f:
    #    f.write(html_report)
    #
    # html_report = create_county_report('SD', 'html')
    # with open('san_diego_report.html', 'w', encoding='utf-8') as f:
    #    f.write(html_report)
    html_report = create_county_report('SB', 'html')
    with open('santa_barbara_report.html', 'w', encoding='utf-8') as f:
        f.write(html_report)
