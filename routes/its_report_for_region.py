
import numpy as np 
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import psycopg2
from sqlalchemy import create_engine, text
from jinja2 import Template
import pdfkit
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import base64
import os
import re
from datetime import datetime
from collections import defaultdict
import geopandas as gpd
import plotly.express as px
import json
import shapely.wkt
import shapely.geometry
import io
from io import BytesIO
import markdown
from dotenv.main import load_dotenv
from openai import OpenAI
import plotly.graph_objects as go
import logging
import traceback
from fastapi import APIRouter, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, Response
from sqlalchemy import text
from sqlalchemy.orm import Session
from controller.db import SessionLocal, get_db
from routes.crop_annualburnprobability_to_region import create_region_map
import concurrent.futures
from fastapi import Depends

import sys
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
handler = logging.StreamHandler()
logger.addHandler(handler)


def debug_print(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    sys.stdout.write(f"[{timestamp}] {msg}\n")
    sys.stdout.flush()


router = APIRouter(tags=["Utility"], prefix='/Utility')

load_dotenv('fastApi/.env')

client = OpenAI(api_key=os.environ['openai_api_key'])    

# County abbreviation to full name mapping
COUNTY_MAPPING = {
    'ALA': 'Alameda', 'ALP': 'Alpine', 'AMA': 'Amador', 'BUT': 'Butte',
    'CAL': 'Calaveras', 'CC': 'Contra Costa', 'COL': 'Colusa', 'DN': 'Del Norte',
    'ED': 'El Dorado', 'FRE': 'Fresno', 'GLE': 'Glenn', 'HUM': 'Humboldt',
    'IMP': 'Imperial', 'INY': 'Inyo', 'KER': 'Kern', 'KIN': 'Kings',
    'LA': 'Los Angeles', 'LAK': 'Lake', 'LAS': 'Lassen', 'MAD': 'Madera',
    'MEN': 'Mendocino', 'MER': 'Merced', 'MNO': 'Mono', 'MOD': 'Modoc',
    'MON': 'Monterey', 'MPA': 'Mariposa', 'MRN': 'Marin', 'NAP': 'Napa',
    'NEV': 'Nevada', 'ORA': 'Orange', 'PLA': 'Placer', 'PLU': 'Plumas',
    'RIV': 'Riverside', 'SAC': 'Sacramento', 'SB': 'Santa Barbara', 'SBD': 'San Bernardino',
    'SBT': 'San Benito', 'SCL': 'Santa Clara', 'SCR': 'Santa Cruz', 'SD': 'San Diego',
    'SF': 'San Francisco', 'SHA': 'Shasta', 'SIE': 'Sierra', 'SIS': 'Siskiyou',
    'SJ': 'San Joaquin', 'SLO': 'San Luis Obispo', 'SM': 'San Mateo', 'SOL': 'Solano',
    'SON': 'Sonoma', 'STA': 'Stanislaus', 'SUT': 'Sutter', 'TEH': 'Tehama',
    'TRI': 'Trinity', 'TUL': 'Tulare', 'TUO': 'Tuolumne', 'VEN': 'Ventura',
    'YOL': 'Yolo', 'YUB': 'Yuba'
}

# Agency abbreviation to full name mapping
AGENCY_MAPPING = {
    'CALSTA': 'California State Transportation Agency',
    'CA State Parks': 'California State Parks',
    'CNRA': 'California Natural Resources Agency',
    'DOD': 'Department of Defense',
    'DOI': 'Department of the Interior',
    'NPS': 'National Park Service',
    'OTHER': 'Other Agencies',
    'TIMBER': 'Timber Operations',
    'Timber Companies': 'Timber Companies',
    'USDA': 'United States Department of Agriculture',
    'US Department of Energy': 'United States Department of Energy'
}

# Activity category to full name mapping
ACTIVITY_CAT_MAPPING = {
    'GRAZING': 'Grazing',
    'LAND_PROTEC': 'Land Protection',
    'MECH_HFR': 'Mechanical & Hand Fuels Reduction',
    'NOT_DEFINED': 'Not Defined',
    'PRESCRIBED_FIRE': 'Prescribed Fire',
    'TIMB_HARV': 'Timber Harvest',
    'TREE_PLNTING': 'Tree Planting'
}

# Activity description to full name mapping
ACTIVITY_DESC_MAPPING = {
    "AMW_AREA_RESTOR": "Area-Wide Restoration",
    "BIOMASS_REMOVAL": "Biomass Removal",
    "BROADCAST_BURN": "Broadcast Burn",
    "CHAIN_CRUSH": "Chain Crushing",
    "CHIPPING": "Chipping",
    "CLEARCUT": "Clearcutting",
    "COMM_THIN": "Commercial Thinning",
    "CONVERSION": "Land Conversion",
    "DISCING": "Discing",
    "DOZER_LINE": "Dozer Line Construction",
    "EASEMENT": "Easement",
    "ECO_HAB_RESTORATION": "Ecological Habitat Restoration",
    "EROSION_CONTROL": "Erosion Control",
    "FEE_TITLE": "Fee Title Acquisition",
    "GRP_SELECTION_HARVEST": "Group Selection Harvest",
    "HABITAT_REVEG": "Habitat Revegetation",
    "HANDLINE": "Handline Construction",
    "HERBICIDE_APP": "Herbicide Application",
    "INV_PLANT_REMOVAL": "Invasive Plant Removal",
    "LAND_ACQ": "Land Acquisition",
    "LANDING_TRT": "Landing Treatment",
    "LOP_AND_SCAT": "Lop and Scatter",
    'MASTICATION': 'Mastication',
    'MOWING': 'Mowing',
    'NOT_DEFINED': 'Not Defined',
    'OAK_WDLND_MGMT': 'Oak Woodland Management',
    'PEST_CNTRL': 'Pest Control',
    'PILE_BURN': 'Pile Burning',
    'PILING': 'Piling',
    'PL_TREAT_BURNED': 'Planned Treatment (Burned)',
    'PRESCRB_HERBIVORY': 'Prescribed Herbivory',
    'PRUNING': 'Pruning',
    'REHAB_UNDRSTK_AREA': 'Rehabilitation Understocked Area',
    'ROAD_CLEAR': 'Road Clearing',
    'ROAD_OBLITERATION': 'Road Obliteration',
    'SALVG_HARVEST': 'Salvage Harvest',
    'SANI_HARVEST': 'Sanitation Harvest',
    'SEEDBED_PREP': 'Seedbed Preparation',
    'SHELTERWD_REM_STEP': 'Shelterwood Removal Step',
    'SINGLE_TREE_SELECTION': 'Single Tree Selection',
    'SITE_PREP': 'Site Preparation',
    'SLASH_DISPOSAL': 'Slash Disposal',
    'SP_PRODUCTS': 'Special Products',
    'STREAM_CHNL_IMPRV': 'Stream Channel Improvement',
    'TBD': 'To Be Determined',
    'THIN_MAN': 'Thinning (Manual)',
    'THIN_MECH': 'Thinning (Mechanical)',
    'TRANSITION_HARVEST': 'Transition Harvest',
    'TREE_FELL': 'Tree Felling',
    'TREE_PLNTING': 'Tree Planting',
    'TREE_RELEASE_WEED': 'Tree Release/Weeding',
    'TREE_SEEDING': 'Tree Seeding',
    'UTIL_RIGHTOFWAY_CLR': 'Utility Right-of-Way Clearing',
    'VARIABLE_RETEN_HARVEST': 'Variable Retention Harvest',
    'WETLAND_RESTOR': 'Wetland Restoration',
    'WM_RESRC_BENEFIT': 'Watershed Management Resource Benefit',
    'YARDING': 'Yarding'
}

# Vegetation type to full name mapping
VEGETATION_MAPPING = {
    'AGRICULTURE': 'Agricultural Land',
    'FOREST': 'Forest Land',
    'GRASS_HERB': 'Grassland/Herbaceous',
    'SHRB_CHAP': 'Shrub/Chaparral',
    'SPARSE': 'Sparse Vegetation',
    'Trees Removed': 'Trees Removed',
    'URBAN': 'Urban/Developed',
    'WATER': 'Water',
    'WETLAND': 'Wetland'
}

# Ownership group to full name mapping
OWNERSHIP_MAPPING = {
    'FEDERAL': 'Federal Government',
    'LOCAL': 'Local Government',
    'NGO': 'Non-Governmental Organization',
    'PRIVATE_INDUSTRY': 'Private Industry',
    'PRIVATE_NON-INDUSTRY': 'Private Non-Industry',
    'STATE': 'State Government',
    'TRIBAL': 'Tribal Land'
}

# Region to full name mapping
REGION_MAPPING = {
    'CENTRAL_COAST': 'Central Coast',
    'NORTH_COAST': 'North Coast',
    'SIERRA_NEVADA': 'Sierra Nevada',
    'SOUTHERN_CA': 'Southern California',
    'Non-Spatial Data': 'Non-Spatial Data'
}

# Activity status to full name mapping
STATUS_MAPPING = {
    'ACTIVE': 'Active',
    'CANCELLED': 'Cancelled',
    'COMPLETE': 'Complete',
    'PLANNED': 'Planned'
}

# Administering organization to full name mapping
ADMIN_ORG_MAPPING = {
    'BIA': 'Bureau of Indian Affairs',
    'BLM': 'Bureau of Land Management',
    'BOF': 'Board of Forestry',
    'CALFIRE': 'California Department of Forestry and Fire Protection',
    'CALTRANS': 'California Department of Transportation',
    'CCC': 'California Conservation Corps',
    'CDFW': 'California Department of Fish and Wildlife',
    'DOC': 'Department of Conservation',
    'DOD': 'Department of Defense',
    'FWS': 'Fish and Wildlife Service',
    'MRCA': 'Mountains Recreation and Conservation Authority',
    'NPS': 'National Park Service',
    'OTHER': 'Other Organizations',
    'PARKS': 'California State Parks',
    'RMC': 'Rivers and Mountains Conservancy',
    'SCC': 'State Coastal Conservancy',
    'SDRC': 'San Diego River Conservancy',
    'SMMC': 'Santa Monica Mountains Conservancy',
    'SNC': 'Sierra Nevada Conservancy',
    'TAHOE': 'Tahoe Conservancy',
    'TIMBER': 'Timber Operations',
    'US Department of Energy': 'United States Department of Energy',
    'USFS': 'United States Forest Service',
    'WCB': 'Wildlife Conservation Board'
}

BOUNDARY_DATASET_TABLE_MAPPING = {
    'California Counties':                                      ('boundary.ca_counties',                                              'namelsad'),
    'California Local Fire Districts':                          ('boundary.california_local_fire_districts',                          'name'),
    "Regional Resource Kit Boundaries":                         ('boundary.rrk_boundaries',                                           'rrk_region'),
    "BLM CA Administrative Unit Boundary Field Office Polygon": ('boundary.blm_ca_administrative_unit_boundary_field_office_polygon', 'admu_name'),
    'Administrative Forest Boundaries':                         ('boundary.forest_administrative_boundaries',                         'forestname'),
    'CAL FIRE Operational Units':                               ('boundary.cal_fire_operational_units',                               'unit'),
    'California State Senate Districts':                        ('boundary.california_state_senate_districts_map_2020',               'name'),
    'California Assembly Districts':                            ('boundary.assembly_districts',                                       'assemblydi')
}




def create_db_connection():
    try:
        connection_string = os.environ['database_connection']
        engine = create_engine(connection_string)
        return engine
    except Exception as e:
        logger.error(f"Error connecting to database: {e}")
        return None

def get_region_data(db: Session, table_name: str, column_name: str, region_name: str):
    """Retrieve treatment data for a region from any PostGIS table"""
    try:
        # Get treatment data intersecting with the region geometry
        data_query = f"""
           WITH region_geom AS (                                                                                                                          
                    SELECT ST_Transform(geom, 4269) AS geom                                                                                               
                    FROM {table_name}                                                                                                                     
                   WHERE {column_name} = '{region_name}'                                                                                                  
                ),                                                                                                                                        
                bbox AS (                                                                                                                                 
                     SELECT ST_SetSRID(ST_Extent(geom), 4269)::geometry AS geom                                                                           
                     FROM region_geom                                                                                                                     
                )                                                                                                                                         
            SELECT its.*,
                   ST_X(ST_Transform(its.geom, 4326)) as x,
                   ST_Y(ST_Transform(its.geom, 4326)) as y 
            FROM its.activities_report_20250110 AS its,      
                 bbox,                                                                                                                                    
                 region_geom                                                                                                                              
            WHERE ST_Within(its.geom, bbox.geom) -- Now bbox.geom is a proper geometry                                                                    
              AND st_contains(region_geom.geom, its.geom)                                                                                                 
              AND year_txt ~ '^[0-9]+$'                                                                                                                   
              AND CAST(year_txt AS INTEGER) BETWEEN 2021 AND 2023
            """

        print("-"*70)
        print(data_query)

        with db.bind.connect() as conn:
            df = pd.read_sql_query(text(data_query), conn)
        return df
    
    except Exception as e:
        logger.error(f"Error fetching region data: {str(e)}")
        return pd.DataFrame()

# Keep all original chart functions but rename county parameters to region
def plot_to_base64(plt_figure):
    buf = io.BytesIO()
    plt_figure.savefig(buf, format='png', dpi=300, bbox_inches='tight')
    buf.seek(0)
    img_str = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close(plt_figure)
    return img_str

def adjust_chart_scaling(ax, data):
    max_val = data.max().max()
    min_val = data.min().min()
    if min_val == 0:
        min_val = 1e-6
    if False and max_val / min_val > 100:
        ax.set_yscale("log")
        ax.set_ylabel("Treatment Acres (Log Scale)")
    else:
        ax.set_ylabel("Treatment Acres")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

def create_agency_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["agency_name"] = df["agency"].map(AGENCY_MAPPING).fillna(df["agency"])
    agency_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "agency_name"])['activity_quantity'].sum().unstack(fill_value=0)
    agency_year_acres = agency_year_acres.clip(upper=np.percentile(agency_year_acres, 95))
    fig, ax = plt.subplots(figsize=(12, 6))
    agency_year_acres.plot(kind="bar", ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Agency (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, agency_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Agency", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_activity_cat_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["activity_cat_name"] = df["activity_cat"].map(ACTIVITY_CAT_MAPPING).fillna(df["activity_cat"])
    activity_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "activity_cat_name"])['activity_quantity'].sum().unstack(fill_value=0)
    activity_year_acres = activity_year_acres.clip(upper=np.percentile(activity_year_acres, 95))
    fig, ax = plt.subplots(figsize=(10, 6))
    activity_year_acres.plot(kind="bar", ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Category (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, activity_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Category", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_vegetation_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["vegetation_name"] = df["broad_vegetation_type"].map(VEGETATION_MAPPING).fillna(df["broad_vegetation_type"])
    veg_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "vegetation_name"])['activity_quantity'].sum().unstack(fill_value=0)
    veg_year_acres = veg_year_acres.clip(upper=np.percentile(veg_year_acres, 95))
    fig, ax = plt.subplots(figsize=(10, 6))
    veg_year_acres.plot(kind="bar", ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Vegetation Type (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, veg_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Vegetation Type", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_ownership_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["ownership_name"] = df["primary_ownership_group"].map(OWNERSHIP_MAPPING).fillna(df["primary_ownership_group"])
    ownership_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "ownership_name"])['activity_quantity'].sum().unstack(fill_value=0)
    ownership_year_acres = ownership_year_acres.clip(upper=np.percentile(ownership_year_acres, 95))
    fig, ax = plt.subplots(figsize=(10, 6))
    ownership_year_acres.plot(kind="bar", ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Land Ownership (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, ownership_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Ownership", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_status_chart(df):
    plt.style.use("seaborn-v0_8-whitegrid")
    df["status_name"] = df["activity_status"].map(STATUS_MAPPING).fillna(df["activity_status"])
    status_year_acres = df[df['activity_uom'] == 'AC'].groupby(["year_txt", "status_name"])['activity_quantity'].sum().unstack(fill_value=0)
    status_year_acres = status_year_acres.clip(upper=np.percentile(status_year_acres, 95))
    fig, ax = plt.subplots(figsize=(10, 6))
    status_year_acres.plot(kind="bar", stacked=True, ax=ax, width=0.8, colormap="coolwarm")
    ax.set_title("Treatment Acres by Status (2021-2023)", fontsize=14, fontweight="bold", pad=20)
    adjust_chart_scaling(ax, status_year_acres)
    ax.set_xlabel("Year", fontsize=12)
    ax.legend(title="Status", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.xticks(rotation=0)
    plt.tight_layout()
    return fig

def create_admin_org_chart(df):
    plt.style.use('seaborn-v0_8-whitegrid')
    df['admin_org_name'] = df['administering_org'].map(ADMIN_ORG_MAPPING).fillna(df['administering_org'])
    top_orgs = df[df['activity_uom'] == 'AC'].groupby('admin_org_name')['activity_quantity'].sum().nlargest(10)
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = plt.cm.viridis(np.linspace(0.1, 0.9, len(top_orgs)))
    bars = ax.barh(top_orgs.index[::-1], top_orgs.values[::-1], color=colors[::-1], height=0.7, edgecolor='none')
    for i, (value, bar) in enumerate(zip(top_orgs.values[::-1], bars)):
        ax.text(value + (value * 0.02), i, f'{int(value):,}', va='center', ha='left', fontsize=10, fontweight='bold')
    ax.set_title('Top 10 Administering Organizations by Acres', fontsize=14, fontweight='bold', pad=20)
    ax.set_xlabel('Acres Treated', fontsize=12)
    ax.set_ylabel('')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.3)
    plt.tight_layout()
    return fig

def generate_summary_statistics(df, region_name):
    df_filtered = df[df['activity_uom'] == 'AC']
    total_activities = len(df_filtered)
    total_acres = df_filtered['activity_quantity'].sum()
    years_active = sorted(df_filtered[df_filtered['year_txt'].str.isnumeric()]['year_txt'].astype(int).unique())
    top_activity = df_filtered.groupby('activity_description')['activity_quantity'].sum().idxmax()
    top_activity = ACTIVITY_DESC_MAPPING.get(top_activity, top_activity)
    main_agency = df_filtered.groupby('agency')['activity_quantity'].sum().idxmax()
    main_agency = AGENCY_MAPPING.get(main_agency, main_agency)
    return {
        'region_name': region_name,
        'total_activities': total_activities,
        'total_acres': f"{total_acres:,.0f}",
        'years_active': f"{min(years_active)} to {max(years_active)}",
        'top_activity': top_activity,
        'main_agency': main_agency
    }

def create_activity_description_table(df):
    df_filtered = df[df['activity_uom'] == 'AC'].copy()
    df_filtered['activity_desc_name'] = df_filtered['activity_description'].map(ACTIVITY_DESC_MAPPING).fillna(df_filtered['activity_description'])
    activity_acres = df_filtered.groupby('activity_desc_name')['activity_quantity'].sum().reset_index()
    activity_acres.columns = ['Treatment Activity', 'Total Acres Treated']
    total_acres = activity_acres['Total Acres Treated'].sum()
    activity_acres['Percentage'] = round(activity_acres['Total Acres Treated'] / total_acres * 100, 1)
    activity_acres['Percentage'] = activity_acres['Percentage'].astype(str) + '%'
    activity_acres['Total Acres Treated'] = activity_acres['Total Acres Treated'].map(lambda x: f"{x:,.2f}")
    html_table = activity_acres.to_html(index=False, classes='table table-striped table-hover', border=0)
    return html_table



def extract_data_insights(df, county_name):
    df_filtered = df[df['activity_uom'] == 'AC'].copy()
    yearly_totals = df_filtered.groupby('year_txt')['activity_quantity'].sum()
    try:
        peak_year = yearly_totals.idxmax()
        peak_amount = yearly_totals.max()
    except:
        peak_year = "N/A"
        peak_amount = 0
    if len(yearly_totals) > 1:
        years = sorted(yearly_totals.index)
        first_year = years[0]
        last_year = years[-1]
        first_amount = yearly_totals[first_year]
        last_amount = yearly_totals[last_year]
        percent_change = ((last_amount - first_amount) / first_amount * 100) if first_amount > 0 else 0
        trend_direction = "increased" if percent_change > 0 else "decreased"
    else:
        percent_change = 0
        trend_direction = "remained stable"
    top_activities = df_filtered.groupby('activity_description')['activity_quantity'].sum().nlargest(3)
    top_activities_names = [ACTIVITY_DESC_MAPPING.get(act, act) for act in top_activities.index]
    top_veg_types = df_filtered.groupby('broad_vegetation_type')['activity_quantity'].sum().nlargest(3)
    top_veg_names = [VEGETATION_MAPPING.get(veg, veg) for veg in top_veg_types.index]
    top_ownership = df_filtered.groupby('primary_ownership_group')['activity_quantity'].sum().nlargest(3)
    top_ownership_names = [OWNERSHIP_MAPPING.get(own, own) for own in top_ownership.index]
    status_dist = df_filtered.groupby('activity_status')['activity_quantity'].sum()
    total_acres = status_dist.sum()
    complete_pct = (status_dist.get('COMPLETE', 0) / total_acres * 100) if total_acres > 0 else 0
    planned_pct = (status_dist.get('PLANNED', 0) / total_acres * 100) if total_acres > 0 else 0
    agency_contrib = df_filtered.groupby('agency')['activity_quantity'].sum().nlargest(2)
    top_agencies = [AGENCY_MAPPING.get(ag, ag) for ag in agency_contrib.index]
    insights = {
        "county_name": county_name,
        "total_acres_treated": df_filtered['activity_quantity'].sum(),
        "peak_year": peak_year,
        "peak_amount": peak_amount,
        "trend_direction": trend_direction,
        "percent_change": round(abs(percent_change), 1),
        "top_activities": top_activities_names,
        "top_vegetation_types": top_veg_names,
        "top_ownership_types": top_ownership_names,
        "completed_treatments_percent": round(complete_pct, 1),
        "planned_treatments_percent": round(planned_pct, 1),
        "lead_agencies": top_agencies
    }
    return insights


def create_admin_org_table(df):
    """Create styled table for administering org data using activity categories"""
    admin_df = df[df['activity_uom'] == 'AC'].groupby(
        ['administering_org', 'activity_cat', 'year_txt']
    )['activity_quantity'].sum().reset_index()
    
    admin_df['Administering Organization'] = admin_df['administering_org'].map(ADMIN_ORG_MAPPING)
    admin_df['Activity Category'] = admin_df['activity_cat'].map(ACTIVITY_CAT_MAPPING)
    admin_df['Year'] = admin_df['year_txt']
    admin_df['Acres'] = admin_df['activity_quantity'].apply(lambda x: f"{x:,.0f}")
    
    return admin_df[['Administering Organization', 'Activity Category', 'Year', 'Acres']] \
        .sort_values(['Year', 'Administering Organization']) \
        .to_html(index=False, classes='table table-striped table-hover', border=0)

def create_ownership_table(df):
    """Create styled table for ownership data using activity categories"""
    ownership_df = df[df['activity_uom'] == 'AC'].groupby(
        ['primary_ownership_group', 'activity_cat', 'year_txt']
    )['activity_quantity'].sum().reset_index()
    
    ownership_df['Ownership'] = ownership_df['primary_ownership_group'].map(OWNERSHIP_MAPPING)
    ownership_df['Activity Category'] = ownership_df['activity_cat'].map(ACTIVITY_CAT_MAPPING)
    ownership_df['Year'] = ownership_df['year_txt']
    ownership_df['Acres'] = ownership_df['activity_quantity'].apply(lambda x: f"{x:,.0f}")
    
    return ownership_df[['Ownership', 'Activity Category', 'Year', 'Acres']] \
        .sort_values(['Year', 'Ownership']) \
        .to_html(index=False, classes='table table-striped table-hover', border=0)


def generate_excel_data(df, region_name):
    """Generate Excel data with proper multi-column sorting"""
    from openpyxl.utils import get_column_letter
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Helper function for column width adjustment
        def auto_adjust_columns(worksheet):
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # 1. Administering Organizations (sorted: Org → Category → Year)
        admin_df = df[df['activity_uom'] == 'AC'].groupby(
            ['administering_org', 'activity_cat', 'year_txt']
        )['activity_quantity'].sum().reset_index()
        
        admin_df['Administering Organization'] = admin_df['administering_org'].map(ADMIN_ORG_MAPPING)
        admin_df['Activity Category'] = admin_df['activity_cat'].map(ACTIVITY_CAT_MAPPING)
        admin_df['Year'] = admin_df['year_txt'].astype(int)
        admin_df['Acres'] = admin_df['activity_quantity']
        
        admin_df = admin_df[['Administering Organization', 'Activity Category', 'Year', 'Acres']] \
            .sort_values(['Administering Organization', 'Activity Category', 'Year'])  # Changed sort order
        
        admin_df.to_excel(writer, sheet_name='Administering Organizations', index=False)
        auto_adjust_columns(writer.sheets['Administering Organizations'])

        # 2. Land Ownership (sorted: Ownership → Category → Year)
        ownership_df = df[df['activity_uom'] == 'AC'].groupby(
            ['primary_ownership_group', 'activity_cat', 'year_txt']
        )['activity_quantity'].sum().reset_index()
        
        ownership_df['Ownership'] = ownership_df['primary_ownership_group'].map(OWNERSHIP_MAPPING)
        ownership_df['Activity Category'] = ownership_df['activity_cat'].map(ACTIVITY_CAT_MAPPING)
        ownership_df['Year'] = ownership_df['year_txt'].astype(int)
        ownership_df['Acres'] = ownership_df['activity_quantity']
        
        ownership_df = ownership_df[['Ownership', 'Activity Category', 'Year', 'Acres']] \
            .sort_values(['Ownership', 'Activity Category', 'Year'])  # Changed sort order
        
        ownership_df.to_excel(writer, sheet_name='Land Ownership', index=False)
        auto_adjust_columns(writer.sheets['Land Ownership'])

        # Format numeric columns
        for sheet in writer.sheets.values():
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 4:  # Acres column
                        cell.number_format = '#,##0.00'
                    elif cell.column == 3:  # Year column
                        cell.number_format = '0'

    output.seek(0)
    return output



@router.get("/its_region_report", include_in_schema=True)
def create_region_report(
    boundary_dataset_name: str,
    region_name: str,
    output_format: str = 'html',
    db: Session = Depends(get_db)
):

    if boundary_dataset_name in BOUNDARY_DATASET_TABLE_MAPPING.keys():
        table_name = BOUNDARY_DATASET_TABLE_MAPPING[boundary_dataset_name][0]
        column_name = BOUNDARY_DATASET_TABLE_MAPPING[boundary_dataset_name][1]    
    else:
        raise HTTPException(status_code=404,                                                                                                          
                detail=f"No boundary dataset with the name {boundary_dataset_name}")    


    # Get treatment data
    debug_print("Getting region points")
    df = get_region_data(db, table_name, column_name, region_name)
    if df.empty:
        raise HTTPException(status_code=404, 
            detail=f"No treatment data found for {region_name} in {boundary_dataset_name}")
    debug_print(f"Got region points: {df.shape}")

    # Add Excel output handling
    if output_format.lower() == 'xlsx':
        excel_data = generate_excel_data(df, region_name)
        return StreamingResponse(
            excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={region_name}_treatment_report.xlsx"}
        )

    try:        
        insights = extract_data_insights(df, region_name)
        agency_chart = plot_to_base64(create_agency_chart(df))
        activity_cat_chart = plot_to_base64(create_activity_cat_chart(df))
        vegetation_chart = plot_to_base64(create_vegetation_chart(df))
        ownership_chart = plot_to_base64(create_ownership_chart(df))
        status_chart = plot_to_base64(create_status_chart(df))
        admin_org_chart = plot_to_base64(create_admin_org_chart(df))
        summary = generate_summary_statistics(df, region_name)
        activity_table = create_activity_description_table(df)

        # Generate region map
        try:
            debug_print(f"Generating overlay_map")
            overlay_map = create_region_map(
                db=db,
                points_df=df,
                table_name=table_name,
                column_name=column_name,
                region_name=region_name,
                geoserver_url="https://sparcal.sdsc.edu/geoserver",
                layer_name="rrk:annualburnprobability_202212_202406_t1_v5"
                # layer_name="rrk:wildlifespecrichness_202304_202406_t1_v5"
            )
            debug_print(f"Generated: overlay_map")
        except Exception as e:
            traceback.print_exc()
            logger.error(f"Map generation error: {str(e)}")
            overlay_map = None
    
        logger.info("Generating spreadsheet tables")
        admin_table = create_admin_org_table(df)
        ownership_table = create_ownership_table(df)
        
        # Generate HTML template
        html_template = Template('''
<!DOCTYPE html>
<html>
<head>
    <title>{{ summary.region_name }} Wildfire Report</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        :root {
            --primary-color: #1e5c97;
            --secondary-color: #2c8b57;
            --accent-color: #f39c12;
            --light-bg: #f8f9fa;
            --dark-bg: #2c3e50;
            --text-color: #333333;
            --light-text: #ffffff;
            --border-color: #e0e0e0;
            --box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }
        body {
            font-family: 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            line-height: 1.6;
            color: var(--text-color);
            max-width: 1200px;
            margin: 0 auto;
            padding: 0;
            background-color: #f5f7fa;
        }
        h1, h2, h3, h4 {
            font-weight: 600;
            color: var(--primary-color);
            margin-top: 1.5em;
            margin-bottom: 0.8em;
        }
        h1 {
            font-size: 2.5rem;
            margin-top: 0;
        }
        h2 {
            font-size: 1.8rem;
            border-bottom: 2px solid var(--border-color);
            padding-bottom: 0.3em;
        }
        h3 {
            font-size: 1.4rem;
            color: var(--secondary-color);
        }
        .container {
            background-color: white;
            box-shadow: var(--box-shadow);
            padding: 2rem;
            margin: 0 auto;
        }
        .header {
            background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
            color: var(--light-text);
            padding: 3rem;
            margin-bottom: 2rem;
            text-align: center;
            position: relative;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
        }
        .header h1, .header h2 {
            color: white;
            margin: 0.5rem 0;
            text-shadow: 1px 1px 3px rgba(0, 0, 0, 0.2);
        }
        .header p {
            margin-top: 1rem;
            font-size: 1.1rem;
            opacity: 0.9;
        }
        .section {
            margin-bottom: 2.5rem;
            padding: 0 1.5rem;
        }
        .introduction {
            background-color: var(--light-bg);
            border-left: 4px solid var(--primary-color);
            padding: 1.5rem;
            margin-bottom: 2rem;
            border-radius: 0 4px 4px 0;
        }
        .summary-box {
            background-color: white;
            border-radius: 8px;
            box-shadow: var(--box-shadow);
            padding: 1.5rem 2rem;
            margin-bottom: 2rem;
            border-top: 5px solid var(--accent-color);
        }
        .summary-box h2 {
            color: var(--accent-color);
            border-bottom: none;
            margin-top: 0;
        }
        .summary-stats {
            display: flex;
            flex-wrap: wrap;
            gap: 1.5rem;
            margin-top: 1.5rem;
        }
        .stat-item {
            flex: 1;
            min-width: 200px;
            background-color: var(--light-bg);
            padding: 1rem;
            border-radius: 6px;
            text-align: center;
        }
        .stat-value {
            font-size: 1.8rem;
            font-weight: 700;
            color: var(--primary-color);
            margin-bottom: 0.5rem;
        }
        .stat-label {
            font-size: 0.9rem;
            color: #666;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .chart-container {
            margin: 2rem 0;
        }
        .chart-row {
            display: flex;
            flex-wrap: wrap;
            gap: 2rem;
            margin-bottom: 2rem;
        }
        .chart {
            flex: 1;
            min-width: 300px;
            background-color: white;
            border-radius: 8px;
            box-shadow: var(--box-shadow);
            padding: 1.5rem;
            transition: transform 0.2s ease;
        }
        .chart:hover {
            transform: translateY(-5px);
        }
        .chart h3 {
            text-align: center;
            margin-top: 0;
            padding-bottom: 0.8rem;
            border-bottom: 1px solid var(--border-color);
        }
        .chart img {
            max-width: 100%;
            height: auto;
            display: block;
            margin: 1rem auto;
        }
        .chart-caption {
            font-size: 0.9rem;
            color: #666;
            text-align: center;
            margin-top: 1rem;
            font-style: italic;
        }
        .full-width {
            flex-basis: 100%;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin: 2rem 0;
            background-color: white;
            box-shadow: var(--box-shadow);
            border-radius: 8px;
            overflow: hidden;
        }
        th, td {
            padding: 12px 15px;
            text-align: left;
        }
        th {
            background-color: var(--primary-color);
            color: white;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.9rem;
            letter-spacing: 0.5px;
        }
        tr:nth-child(even) {
            background-color: #f2f7ff;
        }
        tr:hover {
            background-color: #e6f0ff;
        }
        .findings-box {
            background-color: var(--light-bg);
            padding: 1.5rem 2rem;
            border-radius: 8px;
            margin-bottom: 2rem;
            border-left: 4px solid var(--primary-color);
        }
        .recommendations {
            background-color: var(--light-bg);
            padding: 1.5rem 2rem;
            border-radius: 8px;
            border-left: 4px solid var(--secondary-color);
        }
        .recommendations ul {
            padding-left: 1.2rem;
        }
        .recommendations li {
            margin-bottom: 0.8rem;
        }
        .footer {
            margin-top: 3rem;
            padding: 2rem 0;
            text-align: center;
            font-size: 0.9rem;
            color: #777;
            border-top: 1px solid var(--border-color);
        }
        @media print {
            body {
                background-color: white;
            }
            .container {
                box-shadow: none;
                padding: 0;
            }
            .chart:hover {
                transform: none;
            }
            .chart-row {
                display: block;
            }
            .chart {
                width: 100%;
                margin-bottom: 2rem;
                box-shadow: none;
                page-break-inside: avoid;
            }
            .header {
                background: var(--primary-color) !important;
                -webkit-print-color-adjust: exact;
            }
            th {
                background-color: var(--primary-color) !important;
                color: white !important;
                -webkit-print-color-adjust: exact;
            }
            .summary-box {
                box-shadow: none;
                border: 1px solid var(--border-color);
            }
            table {
                box-shadow: none;
            }

    /* Chart sizing adjustments */
    .chart img {
        max-height: 320px !important;  /* Increased from 250px */
        width: auto !important;
        margin: 12px auto !important;
        page-break-inside: avoid;
    }

    /* Container adjustments */
    .chart {
        page-break-inside: avoid;
        margin: 8px 0 !important;
        padding: 4px !important;
    }

    /* Grid layout for charts */
    .chart-row {
        display: grid !important;
        grid-template-columns: 1fr 1fr;
        gap: 8px !important;
        page-break-inside: avoid;
    }

    /* Single chart full width */
    .full-width.chart img {
        max-height: 400px !important;
        width: 95% !important;
    }

        }


/* Chart image quality preservation */
.chart img {
    image-rendering: crisp-edges;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
}

/* PDF-specific image scaling */
@media print {
    canvas {
        max-width: 100% !important;
        height: auto !important;
    }
    
    figure {
        max-width: 90% !important;
        margin: 0 auto !important;
    }
}


    </style>

</head>
<body>
    <div class="container">
        <div class="header">
            <p>Wildfire & Landscape Resilience Task Force</p>
            <h2>Interagency Treatment Summary Report</h2>
            <p>
                <span style="font-size:16pt; font-weight: bold;">{{ summary.region_name }}</span> 
                <br/>
                {{ boundary_dataset_name }}
            </p>
        </div>
        
        <div class="section">
             <div class="summary-box">
                <h2>Region Treatment Overview</h2>
                <div class="summary-stats">
                    <div class="stat-item">
                        <div class="stat-value">{{ summary.total_activities }}</div>
                        <div class="stat-label">Total Treatment Activities</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value">{{ summary.total_acres }}</div>
                        <div class="stat-label">Total Acres Treated</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value">{{ summary.years_active }}</div>
                        <div class="stat-label">Active Treatment Period</div>
                    </div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <div class="chart-container">
               <div class="chart-row">
                    <div class="chart full-width">
                        <h3>Treatment Locations and Annual Burn Probability</h3>
                        {% if overlay_map %}
                            <img src="data:image/png;base64,{{ overlay_map }}" alt="Burn Probability Overlay">
                        {% else %}
                            <p>Overlay map could not be generated.</p>
                        {% endif %}
			<p class="table-description">
			    Annual Burn Probability - likelihood of a wildfire of any intensity occurring at a given location in a single fire season, <a href="https://caregionalresourcekits.org/clm.html#fire_dyn" target="_blank">California Landscape Metrics</a>, Pyrologix LLC, 2022
			</p>
                         
                   </div>
                </div>
             </div>
        </div>


        <div class="section">
            <h2>Treatment Activities Analysis</h2>
            <div class="chart-container">
                <div class="chart-row">
                    <div class="chart">
                        <h3>Treatment Categories Distribution</h3>
                        <img src="data:image/png;base64,{{ activity_cat_chart }}" alt="Treatment Categories">
                        <p class="chart-caption">Breakdown of treatment activities by category across 2021-2023.</p>
                    </div>
                    <div class="chart">
                        <h3>Vegetation Types Treated</h3>
                        <img src="data:image/png;base64,{{ vegetation_chart }}" alt="Vegetation Types Treated">
                        <p class="chart-caption">Distribution of treatments across vegetation types, 2021-2023.</p>
                    </div>
                </div>
                <div class="chart-row">
                    <div class="chart">
                        <h3>Land Ownership Distribution</h3>
                        <img src="data:image/png;base64,{{ ownership_chart }}" alt="Land Ownership">
                        <p class="chart-caption">Treatment distribution by land ownership, 2021-2023.</p>
                    </div>
                    <div class="chart">
                        <h3>Top Administering Organizations</h3>
                        <img src="data:image/png;base64,{{ admin_org_chart }}" alt="Top Administering Organizations">
                        <p class="chart-caption">Top 10 organizations administering treatments, 2021-2023.</p>
                    </div>
                </div>
            </div>
        </div>

        <div class="section">
            <h2>Detailed Treatment Data</h2>

             <h3>Administering Organizations</h3>
             <p class="table-description">
                   Breakdown of treatment acres by administering organization and activity type.
                   Data reflects actual completed treatments from 2021-2023.
             </p>
             {{ admin_table|safe }}

             <h3>Land Ownership Details</h3>
             <p class="table-description">
                   Distribution of treatments across land ownership types, showing annual 
                   implementation rates by activity category.
             </p>
             {{ ownership_table|safe }}

    </div>

        </div>

        
        <div class="footer">
            <p><strong>California Wildfire and Landscape Interagency Treatment Dashboard</strong></p>
            <p>Report Generated: {{ current_date }}</p>
        </div>
    </div>
</body>
</html>
        ''')
        
        # Render HTML
        html_content = html_template.render(
            boundary_dataset_name=boundary_dataset_name,
            summary=summary,
            overlay_map=overlay_map,
            activity_cat_chart=activity_cat_chart,
            vegetation_chart=vegetation_chart,
            ownership_chart=ownership_chart,
            admin_org_chart=admin_org_chart,
            admin_table=admin_table,
            ownership_table=ownership_table,
            current_date=datetime.now().strftime('%B %d, %Y')
        )
        
        # Handle output format
        if output_format.lower() == 'html':
            return HTMLResponse(content=html_content)
        elif output_format.lower() == 'pdf':
            options = {
                'page-size': 'Letter',
                'margin-top': '0.5in',
                'margin-right': '0.5in',
                'margin-bottom': '0.5in',
                'margin-left': '0.5in',
                'encoding': "UTF-8",
                'enable-local-file-access': None
            }
            pdf = pdfkit.from_string(html_content, False, options=options)
            return Response(content=pdf, media_type="application/pdf")
            
        return HTTPException(status_code=400, detail="Invalid output format")
        
    except Exception as e:
        traceback.print_exc()
        logger.error(f"Report generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


